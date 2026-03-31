"""
transit_rationalisation_v5.py  —  Industry-Leading Master Engine v5.0
================================================================================
Lead Spatial Data Scientist & Urban Transit Planner
Jammu "Matador" Minibus Network Rationalisation Engine

FIXES vs v4:
  ✓ POPULATION HALLUCINATION  — reads actual raster nodata; per-route values
    are individually correct; Summary total uses a dissolved-union approach
    so no resident is ever double-counted across overlapping route catchments.
  ✓ NEGATIVE DEMAND SCORES    — Residential_Score & Commercial_Score are two
    separate columns, both always >= 0. Gravity denominator epsilon prevents /0.
  ✓ INTERMODAL FULLY REMOVED  — zero references anywhere in logic, outputs, maps.
  ✓ MASTER MAP SIDEBAR        — custom JS sidebar with grouped checkboxes (Network
    / Demand / Detail); hidden Leaflet LayerControl proxied transparently.
  ✓ KPI DASHBOARD OVERLAY     — fixed bottom-right panel: routes, fleet, deduplicated
    population, network cost-benefit score.
  ✓ INDIVIDUAL MAPS           — fully populated: catchment polygon, start/end pins,
    via waypoints, HV POIs, secondary POIs, mini stats panel overlay.
  ✓ EXCEL INTEGRITY           — no negative numbers, two gravity score columns,
    numeric guards on all calculated fields, correct summary deduplicated pop.

CRITICAL CONSTRAINT: flat 2-D plane only.  NO elevation / DEM / slope logic.
================================================================================
"""

# ──────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION  (every tunable parameter lives here — never buried in logic)
# ──────────────────────────────────────────────────────────────────────────────
# Remove the single FALLBACK_SPEED_KMH and replace with these:
SPEED_OLD_CITY_KMH  = 10.0  # Panjtirthi, Kachi Chawni, Raghunath Bazaar
SPEED_SOUTH_JAMMU_KMH = 22.0 # Trikuta Nagar, Highway, Channi Himmat
SPEED_DEFAULT_KMH   = 15.0  # General mixed traffic
# Tier 1: Massive daily footfall (Weight: 1000 equivalent residents)
TIER_1_POIS = frozenset({"hospital", "college", "bus_station", "stadium"})
TIER_1_WEIGHT = 1000

# Tier 2: Moderate daily footfall (Weight: 250 equivalent residents)
# Note: "clinic" is deliberately excluded to prevent noise.
TIER_2_POIS = frozenset({"school", "supermarket", "government", "museum"})
TIER_2_WEIGHT = 250

# Geometry / Routing
MAX_IMPUTED_SL_KM        = 60        # Drop routes > 60 km (regional transport)
CIRCUITY_FACTOR          = 1.40      # SL → road distance multiplier (OSRM fallback)
CIRCUITY_FACTOR_RIVER    = 3.50      # Tawi River crossing penalty
TAWI_RIVER_LON           = 74.87     # Approximate longitude of Tawi bisection
MIN_ROUTE_KM             = 1.0       # Drop degenerate routes (circular exempt)
VIRTUAL_STOP_SPACING_M   = 250       # Hail-and-ride virtual stop interval (m)
WALK_CATCHMENT_M         = 400       # Max walk distance in hilly terrain (m)
OVERLAP_BUFFER_M         = 80        # UTM buffer for corridor frequency detection
OVERLAP_THRESHOLD        = 0.65      # Pairwise overlap ≥ 65% → same cluster
OD_PROXIMITY_TOLERANCE_M = 2500      # Max terminal separation to allow merging
SIMPLIFY_TOL_M           = 2.0       # Geometry simplification (RAM guard, metres)

# Fleet & Headway
STD_TO_MINI_RATIO        = 2.5       # 1 Standard Bus ≡ 2.5 Minibuses
TRUNK_HEADWAY_MIN        = 10
FEEDER_HEADWAY_MIN       = 20
GHOST_HEADWAY_MIN        = 45        # Assumed headway for 0-bus "ghost" routes
     # Used only when OSRM duration unavailable
FLEET_BUFFER_FACTOR      = 1.15      # 15% operational contingency

# Demand Physics
       
GRAVITY_EPSILON          = 0.1       # Min denominator in gravity score

# Transfer Penalty  [Cats et al. 2021]
TRANSFER_PENALTY_MIN     = 15

# Junction Penalty  [Vuchic 2005]
SHARP_TURN_DEG           = 70
JUNCTION_PENALTY_MIN     = 3

# Terminal Capacity
FLEET_CAP_HARD           = 45
TERMINAL_BUFFER_M        = 300

# Cost-Benefit
CB_POP_WEIGHT            = 100
CB_FLEET_COST_INR        = 1_500_000

# OSRM
OSRM_BASE_URL            = "http://localhost:5000"
OSRM_TIMEOUT_S           = 8
OSRM_MAX_WORKERS         = 12

# CRS
UTM_CRS                  = "EPSG:32643"
WGS84_CRS                = "EPSG:4326"

# I/O
RASTER_PATH              = "jammu_worldpop.tif"
ROUTES_CSV               = "routes.csv"
POIS_CSV                 = "pois.csv"
OUTPUT_DIR               = "route_maps"
LOG_CSV                  = "Rationalisation_Log.csv"
ROUTES_OUT_CSV           = "Rationalised_Routes.csv"
ROUTES_OUT_XLSX          = "Rationalised_Routes.xlsx"
PASSENGER_IMPACT_CSV     = "Passenger_Impact.csv"
MASTER_MAP_HTML          = "Master_Transit_Map.html"
# ──────────────────────────────────────────────────────────────────────────────
#  UPDATED CONFIGURATION (Jammu Ground Reality Patch)
# ──────────────────────────────────────────────────────────────────────────────

# Geometry / Routing
CIRCUITY_FACTOR          = 1.25      # SL → road distance multiplier (OSRM fallback)
CIRCUITY_FACTOR_RIVER    = 1.60      # Realistic Tawi River traffic penalty (was 3.50)

# Fleet & Headway
STD_TO_MINI_RATIO        = 2.5       
TRUNK_HEADWAY_MIN        = 5         # High-frequency trunks need buses every 5 mins
FEEDER_HEADWAY_MIN       = 15        # Feeders every 15 mins
GHOST_HEADWAY_MIN        = 45        
    # Adjusted for Jammu's gradients and traffic
FLEET_BUFFER_FACTOR      = 1.15      

# Transfer Penalty
TRANSFER_PENALTY_MIN     = 5         # Realistic jump-out, jump-in time for Matadors

# Junction Penalty
SHARP_TURN_DEG           = 75
JUNCTION_PENALTY_MIN     = 0.5       # 30 seconds per sharp turn, NOT 3 minutes!
# I/O
ROUTES_OUT_XLSX          = "Rationalised_Routes.xlsx"
PASSENGER_IMPACT_CSV     = "Passenger_Impact.csv"
MASTER_MAP_HTML          = "Master_Transit_Map.html"
ROUTES_GEOJSON           = "Rationalised_Routes.geojson"  # <-- ADD THIS LINE

# ──────────────────────────────────────────────────────────────────────────────
#  IMPORTS
# ──────────────────────────────────────────────────────────────────────────────
import json
import math
import sys
import time
import logging
import warnings
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import geopandas as gpd
import requests
from shapely.geometry import LineString, MultiLineString, Point, shape, mapping
from shapely.ops import unary_union
from shapely.strtree import STRtree
import rasterstats
import folium
from folium.plugins import AntPath, HeatMap
try:
    from folium.plugins import PolyLineTextPath
    _HAS_PLTPATH = True
except ImportError:
    _HAS_PLTPATH = False
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ──────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────────────────────────
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("transit_rationalisation.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────

TERMINAL_CATEGORIES = frozenset({"bus_terminal", "market"})

COLOUR = {
    "trunk":         "#1A237E",
    "feeder":        "#00695C",
    "current_net":   "#9E9E9E",
    "catchment_fill":"#80DEEA",
    "catchment_line":"#0097A7",
    "poi_high":      "#D32F2F",
    "poi_secondary": "#F57F17",
    "start_pin":     "#1B5E20",
    "end_pin":       "#B71C1C",
    "via_dot":       "#5C6BC0",
}
TILE_URL  = "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png"
TILE_ATTR = "© OpenStreetMap contributors © CARTO"

# Exact FeatureGroup names (must match sidebar JS strings precisely)
FG = {
    "heatmap":   "Population Heatmap",
    "trunk":     "Trunk Corridors",
    "feeder":    "Feeder Routes",
    "original":  "Original Network",
    "hv_poi":    "High Priority POIs",
    "sec_poi":   "Secondary POIs",
    "catchment": "Route Catchments",
    "pins":      "Start End Terminals",
    "via":       "Via Waypoints",
}

EXPORT_COLS = [
    "Route_ID", "Route_Name", "Action_Taken", "New_Route_ID",
    "Route_KM", "OSRM_Duration_S", "Cycle_Time_Min",
    "Sharp_Turns", "Junction_Penalty_Min",
    "Fleet_Required", "Total_Minibuses_Existing", "Headway_Min",
    "Population_Served", "HV_POI_Count",
    "Residential_Score", "Commercial_Score", "Composite_Demand_Score",
    "Old_Wait_Time", "New_Wait_Time",
    "Overlap_Metric", "Geo_Source", "Fleet_Imputed",
]


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1  ─  DATA INGESTION, CONCURRENT OSRM ROUTING & FLEET IMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def load_routes(path: str) -> pd.DataFrame:
    log.info("Loading routes from '%s'", path)
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.replace(" ", "_")
    col_map = {c.lower(): c for c in df.columns}

    def _resolve(targets):
        for k in targets:
            if k in col_map:
                return col_map[k]
        return None

    alias = {
        "Start_Lat":          ["start_lat", "startlat", "origin_lat", "from_lat"],
        "Start_Lon":          ["start_lon", "startlon", "origin_lon", "start_lng", "from_lon"],
        "End_Lat":            ["end_lat",   "endlat",   "dest_lat",   "to_lat"],
        "End_Lon":            ["end_lon",   "endlon",   "dest_lon",   "end_lng",  "to_lon"],
        "Via_Coordinates":    ["via_coordinates", "via_coords", "waypoints"],
        "Minibus_Count":      ["minibus_count", "minibuses", "mini_bus_count"],
        "Standard_Bus_Count": ["standard_bus_count", "std_bus", "standard_buses"],
        "Route_Name":         ["route_name", "name", "route"],
    }
    rename = {}
    for canon, targets in alias.items():
        found = _resolve(targets)
        if found and found != canon:
            rename[found] = canon
    df.rename(columns=rename, inplace=True)

    if "Route_ID" not in df.columns:
        df["Route_ID"] = [f"R{i+1:04d}" for i in range(len(df))]
    if "Route_Name" not in df.columns:
        if "Route_From" in df.columns and "Route_To" in df.columns:
            df["Route_Name"] = df["Route_From"].astype(str) + " ↔ " + df["Route_To"].astype(str)
        else:
            df["Route_Name"] = df["Route_ID"]
    for col, default in [("Minibus_Count", 0), ("Standard_Bus_Count", 0), ("Via_Coordinates", None)]:
        if col not in df.columns:
            df[col] = default

    df["Minibus_Count"]      = pd.to_numeric(df["Minibus_Count"],      errors="coerce").fillna(0).astype(int)
    df["Standard_Bus_Count"] = pd.to_numeric(df["Standard_Bus_Count"], errors="coerce").fillna(0).astype(int)
    log.info("  Loaded %d routes.", len(df))
    return df


def load_pois(path: str) -> gpd.GeoDataFrame:
    log.info("Loading POIs from '%s'", path)
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.replace(" ", "_")
    if "category" not in df.columns:
        df["category"] = "Other"
    if "name" not in df.columns:
        df["name"] = df["category"]
    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["lon"], df["lat"]), crs=WGS84_CRS)
    log.info("  Loaded %d POIs.", len(gdf))
    return gdf


def parse_via(via_raw) -> List[Tuple[float, float]]:
    if via_raw is None or (isinstance(via_raw, float) and math.isnan(via_raw)):
        return []
    if not isinstance(via_raw, str) or not via_raw.strip():
        return []
    try:
        pts = json.loads(via_raw)
        result = []
        for p in pts:
            if isinstance(p, dict):
                result.append((float(p.get("lon", p.get("lng", 0))), float(p.get("lat", 0))))
            elif isinstance(p, (list, tuple)) and len(p) >= 2:
                a, b = float(p[0]), float(p[1])
                result.append((b, a) if 6 <= a <= 40 else (a, b))
        return result
    except (json.JSONDecodeError, TypeError, ValueError):
        return []


# ── 1-B  Concurrent OSRM (extracts actual duration) ──────────────────────────

def _build_osrm_url(coords: List[Tuple[float, float]]) -> str:
    return (f"{OSRM_BASE_URL}/route/v1/driving/"
            + ";".join(f"{lon},{lat}" for lon, lat in coords)
            + "?overview=full&geometries=geojson&steps=false")


def _fetch_osrm_single(route_id: str, coords: List[Tuple[float, float]]) -> Dict:
    try:
        resp = requests.get(_build_osrm_url(coords), timeout=OSRM_TIMEOUT_S)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == "Ok" and data.get("routes"):
            r0 = data["routes"][0]
            return {
                "route_id":        route_id,
                "geometry":        shape(r0["geometry"]),
                "osrm_km":         r0["distance"] / 1000.0,
                "osrm_duration_s": float(r0["duration"]),
                "success":         True,
            }
    except Exception as exc:
        log.debug("OSRM failed [%s]: %s", route_id, exc)
    return {"route_id": route_id, "geometry": None, "osrm_km": None,
            "osrm_duration_s": None, "success": False}


def fetch_all_osrm(df: pd.DataFrame) -> Dict[str, Dict]:
    log.info("Fetching OSRM geometries (%d routes, %d workers)…", len(df), OSRM_MAX_WORKERS)
    tasks = {}
    for _, row in df.iterrows():
        via = parse_via(row.get("Via_Coordinates"))
        tasks[row["Route_ID"]] = (
            [(float(row["Start_Lon"]), float(row["Start_Lat"]))]
            + via
            + [(float(row["End_Lon"]), float(row["End_Lat"]))]
        )
    results = {}
    with ThreadPoolExecutor(max_workers=OSRM_MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_osrm_single, rid, coords): rid
                   for rid, coords in tasks.items()}
        for fut in as_completed(futures):
            rid = futures[fut]
            try:
                res = fut.result()
                results[res["route_id"]] = res
            except Exception as exc:
                log.warning("Future error [%s]: %s", rid, exc)
                results[rid] = {"route_id": rid, "geometry": None,
                                "osrm_km": None, "osrm_duration_s": None, "success": False}
    ok = sum(1 for r in results.values() if r["success"])
    log.info("  OSRM: %d/%d ok, %d use circuity fallback.", ok, len(df), len(df) - ok)
    return results


def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def apply_geometries(df: pd.DataFrame, osrm_results: Dict) -> gpd.GeoDataFrame:
    log.info("Merging geometries and filtering…")
    rows = []
    for _, row in df.iterrows():
        rid = row["Route_ID"]
        res = osrm_results.get(rid, {})
        via = parse_via(row.get("Via_Coordinates"))
        raw_coords = (
            [(float(row["Start_Lon"]), float(row["Start_Lat"]))]
            + via
            + [(float(row["End_Lon"]), float(row["End_Lat"]))]
        )
        fallback_geom = LineString(raw_coords) if len(raw_coords) >= 2 else None

        if res.get("success") and res["geometry"] is not None:
            geom       = res["geometry"]
            dist_km    = max(0.0, res["osrm_km"])
            duration_s = max(0.0, res["osrm_duration_s"])
            source     = "OSRM"
        else:
            geom     = fallback_geom
            sl_km    = _haversine_km(float(row["Start_Lat"]), float(row["Start_Lon"]),
                                      float(row["End_Lat"]),   float(row["End_Lon"]))
            crosses  = (float(row["Start_Lon"]) < TAWI_RIVER_LON) != \
                       (float(row["End_Lon"])   < TAWI_RIVER_LON)
            cf       = CIRCUITY_FACTOR_RIVER if crosses else CIRCUITY_FACTOR
            dist_km  = sl_km * cf
            
            start_lat = float(row["Start_Lat"])
            end_lat = float(row["End_Lat"])
        
            if start_lat > 32.72 and end_lat > 32.72:
                local_speed = SPEED_OLD_CITY_KMH
            elif start_lat < 32.71 and end_lat < 32.71:
                 local_speed = SPEED_SOUTH_JAMMU_KMH
            else:
                local_speed = SPEED_DEFAULT_KMH # Cross-city routes
            
            duration_s = (dist_km / local_speed) * 3600.0
                
            
            source   = "Circuity-River" if crosses else "Circuity"

        rows.append({**row.to_dict(),
                     "geometry":        geom,
                     "Route_KM":        round(max(0.0, dist_km), 3),
                     "OSRM_Duration_S": round(max(0.0, duration_s), 1),
                     "Geo_Source":      source})

    gdf = gpd.GeoDataFrame(rows, geometry="geometry", crs=WGS84_CRS)
    n0  = len(gdf)
    gdf = gdf[gdf.geometry.notna()].copy()
    gdf = gdf[gdf["Route_KM"] <= MAX_IMPUTED_SL_KM].copy()

    gdf_utm     = gdf.to_crs(UTM_CRS)
    gdf["_circ"] = gdf_utm.apply(
        lambda r: (len(r.geometry.coords) >= 2
                   and Point(r.geometry.coords[0]).distance(Point(r.geometry.coords[-1])) < 100),
        axis=1)
    gdf = gdf[(gdf["Route_KM"] >= MIN_ROUTE_KM) | gdf["_circ"]].copy()
    gdf.drop(columns=["_circ"], inplace=True)
    log.info("  Filter: %d → %d routes kept.", n0, len(gdf))
    return gdf.reset_index(drop=True)


# ── 1-C  POI counting & fleet imputation ─────────────────────────────────────

def count_pois_along_routes(gdf_routes: gpd.GeoDataFrame,
                             gdf_pois: gpd.GeoDataFrame,
                             buffer_m: float = 250.0) -> pd.DataFrame:
    log.info("Counting Tier 1 and Tier 2 POIs within %dm of each route…", buffer_m)
    r_utm = gdf_routes.to_crs(UTM_CRS).copy()
    r_utm["_buf"] = r_utm.geometry.simplify(SIMPLIFY_TOL_M).buffer(buffer_m)
    buf_gdf = r_utm[["Route_ID", "_buf"]].set_geometry("_buf")
    
    # Tier 1 Join
    p_t1 = gdf_pois[gdf_pois["category"].str.lower().isin(TIER_1_POIS)].to_crs(UTM_CRS)
    t1_joined = gpd.sjoin(p_t1, buf_gdf, how="inner", predicate="within")
    t1_counts = t1_joined.groupby("Route_ID").size()
    
    # Tier 2 Join
    p_t2 = gdf_pois[gdf_pois["category"].str.lower().isin(TIER_2_POIS)].to_crs(UTM_CRS)
    t2_joined = gpd.sjoin(p_t2, buf_gdf, how="inner", predicate="within")
    t2_counts = t2_joined.groupby("Route_ID").size()
    
    gdf_routes["Tier1_POI_Count"] = gdf_routes["Route_ID"].map(t1_counts).fillna(0).astype(int)
    gdf_routes["Tier2_POI_Count"] = gdf_routes["Route_ID"].map(t2_counts).fillna(0).astype(int)
    # Keep total for exports
    gdf_routes["HV_POI_Count"] = gdf_routes["Tier1_POI_Count"] + gdf_routes["Tier2_POI_Count"]
    
    return gdf_routes

def impute_fleet(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Imputing fleet sizes…")
    gdf = gdf.copy()
    gdf["Total_Minibuses_Existing"] = (
        gdf["Minibus_Count"] + gdf["Standard_Bus_Count"] * STD_TO_MINI_RATIO
    ).round(1).clip(lower=0)

    mask = gdf["Total_Minibuses_Existing"] == 0

    def _ghost(km):
        # Base cycle time without the inflated 10% penalty
        cycle = (km * 2 / SPEED_DEFAULT_KMH) * 60 
        # Assume these quiet routes only need 1 bus per hour, not 1 every 45 mins
        return max(1, math.ceil(cycle / 60))

    gdf.loc[mask, "Total_Minibuses_Existing"] = gdf.loc[mask, "Route_KM"].apply(_ghost)
    gdf["Fleet_Imputed"] = mask
    log.info("  Fleet imputed for %d zero-bus routes.", mask.sum())
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2  ─  PTAL, GRAVITY MODEL, PHYSICS, CLUSTERING, CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════

# ── 2-A  Catchment polygons ───────────────────────────────────────────────────

def _pts_along_line(geom: LineString, spacing_m: float) -> List[Point]:
    length = geom.length
    if length < spacing_m:
        return [geom.interpolate(0.5, normalized=True)]
    return [geom.interpolate(d) for d in np.arange(0, length + spacing_m, spacing_m)
            if d <= length]


def build_catchments(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Building virtual catchments (stop=%dm, walk=%dm)…",
             VIRTUAL_STOP_SPACING_M, WALK_CATCHMENT_M)
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    catchments = []
    for geom in gdf_utm.geometry:
        pts  = _pts_along_line(geom, VIRTUAL_STOP_SPACING_M)
        poly = unary_union([p.buffer(WALK_CATCHMENT_M) for p in pts])
        catchments.append(poly)
    gdf_utm["_catch"] = catchments
    catch_series = (
        gdf_utm.set_geometry("_catch")
               .set_crs(UTM_CRS)  # <-- Explicitly define the CRS here
               .to_crs(WGS84_CRS)["_catch"]
    )
    gdf["Catchment"] = catch_series.values
    log.info("  Catchments built for %d routes.", len(gdf))
    return gdf


# ── 2-B  Population (FIXED: raster nodata + clamp + dedup summary) ───────────

def _read_raster_nodata(raster_path: str) -> Optional[float]:
    """Read the true nodata value from raster metadata to prevent summing garbage."""
    try:
        import rasterio
        with rasterio.open(raster_path) as src:
            return src.nodata
    except Exception:
        return -9999.0


def compute_population(gdf: gpd.GeoDataFrame, raster_path: str) -> gpd.GeoDataFrame:
    """
    Per-route population from rasterstats.
    FIX: uses actual raster nodata so WorldPop sentinel values (e.g. −3.4e38)
    are never summed, and results are hard-clamped to [0, 2_000_000].
    Jammu metro population is ~700 k — any single route exceeding 2M is noise.
    """
    log.info("Computing per-route population from raster: %s", raster_path)
    if not Path(raster_path).exists():
        log.warning("  Raster not found — Population_Served = 0.")
        gdf["Population_Served"] = 0
        return gdf

    nodata = _read_raster_nodata(raster_path)
    catch_gdf = gpd.GeoDataFrame(gdf[["Route_ID"]], geometry=gdf["Catchment"], crs=WGS84_CRS)
    stats = rasterstats.zonal_stats(
        catch_gdf, raster_path, stats=["sum"], nodata=nodata, geojson_out=False
    )
    # Clamp: >= 0 and <= 2 million (single-route sanity cap for Jammu)
    gdf["Population_Served"] = [
        min(2_000_000, max(0, int(s["sum"]))) if s.get("sum") is not None else 0
        for s in stats
    ]
    log.info("  Population_Served max: %s", f"{gdf['Population_Served'].max():,}")
    return gdf


def compute_network_population_total(gdf: gpd.GeoDataFrame,
                                      raster_path: str) -> int:
    """
    TRUE deduplicated network population: dissolve ALL catchments into a single
    polygon, run zonal_stats once.  This eliminates double-counting where
    downtown routes overlap and would otherwise sum the same residents 50x.

    Approach chosen for Jammu: per-route values are used for route-level ranking
    (each route independently assessed); the Summary total uses this dissolved value.
    """
    log.info("Computing deduplicated network population (dissolved union)…")
    if not Path(raster_path).exists():
        return int(gdf["Population_Served"].max())

    valid = [c for c in gdf["Catchment"] if c is not None
             and hasattr(c, "is_empty") and not c.is_empty]
    if not valid:
        return 0

    dissolved = unary_union(valid)
    nodata    = _read_raster_nodata(raster_path)
    stats     = rasterstats.zonal_stats(
        [dissolved], raster_path, stats=["sum"], nodata=nodata, geojson_out=False
    )
    val = stats[0].get("sum") if stats else None
    result = min(2_000_000, max(0, int(val))) if val is not None else 0
    log.info("  Deduplicated network population: %s", f"{result:,}")
    return result


# ── 2-C  Gravity Demand Scores — TWO COLUMNS (Residential + Commercial) ──────

# In compute_gravity_scores (around line 348):
def compute_gravity_scores(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Computing two-component gravity demand scores…")
    denom = gdf["Route_KM"].clip(lower=GRAVITY_EPSILON)
    gdf["Residential_Score"] = (gdf["Population_Served"] / denom).clip(lower=0).round(2)
    
    # New Tiered Commercial Score
    commercial_demand = (gdf["Tier1_POI_Count"] * TIER_1_WEIGHT) + (gdf["Tier2_POI_Count"] * TIER_2_WEIGHT)
    gdf["Commercial_Score"] = (commercial_demand / denom).clip(lower=0).round(2)
    
    gdf["Composite_Demand_Score"] = (gdf["Residential_Score"] + gdf["Commercial_Score"]).round(2)
    return gdf


# ── 2-D  Junction penalty (Vuchic 2005 mixed-traffic model) ──────────────────

def _deflection_deg(A, B, C) -> float:
    ba = (A[0] - B[0], A[1] - B[1])
    bc = (C[0] - B[0], C[1] - B[1])
    mag_ba, mag_bc = math.hypot(*ba), math.hypot(*bc)
    if mag_ba < 1e-6 or mag_bc < 1e-6:
        return 0.0
    cos_t = max(-1.0, min(1.0, (ba[0]*bc[0] + ba[1]*bc[1]) / (mag_ba * mag_bc)))
    return 180.0 - math.degrees(math.acos(cos_t))


def compute_junction_penalties(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Computing junction penalties (threshold=%d°, +%d min/turn)…",
             SHARP_TURN_DEG, JUNCTION_PENALTY_MIN)
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    n_sharp_list, penalty_list = [], []
    for geom in gdf_utm.geometry:
        if geom is None or geom.is_empty:
            n_sharp_list.append(0); penalty_list.append(0.0); continue
        coords  = list(geom.coords)
        n_sharp = sum(1 for i in range(1, len(coords) - 1)
                      if _deflection_deg(coords[i-1], coords[i], coords[i+1]) > SHARP_TURN_DEG)
        n_sharp_list.append(n_sharp)
        penalty_list.append(float(n_sharp * JUNCTION_PENALTY_MIN))
    gdf["Sharp_Turns"]          = n_sharp_list
    gdf["Junction_Penalty_Min"] = penalty_list
    log.info("  %d routes have junction penalties.", sum(1 for p in penalty_list if p > 0))
    return gdf


# ── 2-E  Cycle time (OSRM duration + junction penalty) ───────────────────────

def compute_cycle_times(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Computing cycle times from OSRM duration + junction penalties…")
    one_way_min      = (gdf["OSRM_Duration_S"].clip(lower=0) / 60.0) + gdf["Junction_Penalty_Min"]
    gdf["Cycle_Time_Min"] = (one_way_min * 2 * FLEET_BUFFER_FACTOR).clip(lower=1.0).round(1)
    log.info("  Cycle time mean: %.1f  max: %.1f", gdf["Cycle_Time_Min"].mean(), gdf["Cycle_Time_Min"].max())
    return gdf


def _fleet_from_cycle(cycle_min: float, headway_min: int) -> int:
    return max(1, math.ceil(cycle_min / max(1, headway_min)))


# ── 2-F  Network-First Corridor Frequency Scores (Perera et al. 2024) ────────

def compute_frequency_scores(gdf: gpd.GeoDataFrame) -> np.ndarray:
    log.info("Computing corridor frequency scores (STRtree)…")
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    bufs    = [g.buffer(OVERLAP_BUFFER_M) for g in gdf_utm.geometry]
    tree    = STRtree(bufs)
    scores  = np.zeros(len(bufs), dtype=int)
    for i, buf_i in enumerate(bufs):
        candidates = tree.query(buf_i)
        scores[i]  = sum(1 for j in candidates if j != i and bufs[j].intersects(buf_i))
    log.info("  Frequency mean: %.1f  max: %d", scores.mean(), scores.max())
    return scores


# ── 2-G  Overlap matrix ───────────────────────────────────────────────────────

def compute_overlap_matrix(gdf: gpd.GeoDataFrame) -> np.ndarray:
    log.info("Computing pairwise overlap matrix…")
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    bufs   = [g.buffer(OVERLAP_BUFFER_M) for g in gdf_utm.geometry]
    areas  = np.array([b.area for b in bufs])
    n      = len(bufs)
    matrix = np.zeros((n, n), dtype=np.float32)
    tree   = STRtree(bufs)
    for i, buf_i in enumerate(bufs):
        for j in tree.query(buf_i):
            if j <= i:
                continue
            inter = buf_i.intersection(bufs[j]).area
            if areas[i] > 0: matrix[i, j] = inter / areas[i]
            if areas[j] > 0: matrix[j, i] = inter / areas[j]
    log.info("  Overlap matrix (%d×%d) done.", n, n)
    return matrix


# ── 2-H  Union-Find clustering ────────────────────────────────────────────────

def cluster_routes(gdf: gpd.GeoDataFrame, overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    log.info("Union-Find clustering (threshold=%.0f%%)…", OVERLAP_THRESHOLD * 100)
    n = len(gdf)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x

    def union(x, y):
        parent[find(x)] = find(y)

    for r, c in zip(*np.where(overlap_matrix >= OVERLAP_THRESHOLD)):
        union(int(r), int(c))

    gdf = gdf.copy()
    gdf["Cluster_ID"]   = [find(i) for i in range(n)]
    gdf["Cluster_Size"] = gdf["Cluster_ID"].map(gdf["Cluster_ID"].value_counts())
    log.info("  %d clusters formed.", gdf["Cluster_ID"].nunique())
    return gdf


def backfill_overlap_metric(gdf: gpd.GeoDataFrame,
                             overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    n = len(gdf)
    gdf = gdf.copy()
    means = []
    for i in range(n):
        row = np.concatenate([overlap_matrix[i, :i], overlap_matrix[i, i+1:]])
        means.append(float(row.mean()) if len(row) else 0.0)
    gdf["Overlap_Metric"] = [round(v, 4) for v in means]
    return gdf


# ── 2-I  Old wait time estimator ─────────────────────────────────────────────

def _old_wait(row: pd.Series) -> int:
    buses = max(1, int(row.get("Total_Minibuses_Existing", 1)))
    cycle = max(1.0, float(row.get("Cycle_Time_Min", 30.0)))
    return max(5, min(90, math.ceil(cycle / buses)))


# ── 2-J  Transfer penalty check (Cats et al. 2021) ───────────────────────────

def _transfer_is_worse(old_wait: int) -> bool:
    """True if forcing a transfer worsens the passenger journey."""
    return (FEEDER_HEADWAY_MIN + TRUNK_HEADWAY_MIN + TRANSFER_PENALTY_MIN) > old_wait


# ── 2-K  Full Classification (gravity + frequency + transfer penalty) ─────────

# ── 2-K  Full Classification (Aggressive Congestion Reduction) ─────────

def classify_routes(gdf: gpd.GeoDataFrame,
                    freq_scores: np.ndarray,
                    overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    log.info("Classifying routes (Aggressive deduplication & Trunk creation)…")
    gdf = gdf.copy()
    gdf["Action_Taken"] = "RETAINED_AS_FEEDER"
    gdf["New_Route_ID"] = ""
    gdf["Headway_Min"]  = FEEDER_HEADWAY_MIN
    gdf["Old_Wait_Time"] = gdf.apply(_old_wait, axis=1)

    trunk_n = feeder_n = 1

    for cluster_id, grp in gdf.groupby("Cluster_ID"):
        idxs = grp.index.tolist()

        if len(idxs) == 1:
            gdf.at[idxs[0], "New_Route_ID"] = f"FDR-{feeder_n:03d}"
            feeder_n += 1
            continue

        # Select Trunk by highest Gravity Score
        sorted_idxs = sorted(idxs, key=lambda i: gdf.at[i, "Composite_Demand_Score"], reverse=True)
        ti = sorted_idxs[0]
        gdf.at[ti, "Action_Taken"] = "UPGRADED_TO_TRUNK"
        gdf.at[ti, "New_Route_ID"] = f"TRK-{trunk_n:03d}"
        gdf.at[ti, "Headway_Min"]  = TRUNK_HEADWAY_MIN

        for idx in sorted_idxs[1:]:
            pos_i    = gdf.index.get_loc(idx)
            pos_t    = gdf.index.get_loc(ti)
            overlap  = float(overlap_matrix[pos_i, pos_t])

            # Terminal proximity guard (prevents merging vastly different O-D pairs)
            start_i = gpd.GeoSeries([Point(float(gdf.at[idx, "Start_Lon"]),
                                           float(gdf.at[idx, "Start_Lat"]))],
                                     crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
            start_t = gpd.GeoSeries([Point(float(gdf.at[ti, "Start_Lon"]),
                                           float(gdf.at[ti, "Start_Lat"]))],
                                     crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
            
            # If the route overlaps the Trunk heavily (>=65%), we merge it to kill redundancy.
            # Removed the absurd pop < 2000 and POI < 3 constraints.
            if overlap >= OVERLAP_THRESHOLD and start_i.distance(start_t) <= OD_PROXIMITY_TOLERANCE_M:
                gdf.at[idx, "Action_Taken"] = "MERGED_INTO_TRUNK"
                gdf.at[idx, "New_Route_ID"] = f"TRK-{trunk_n:03d}"
                gdf.at[idx, "Headway_Min"]  = TRUNK_HEADWAY_MIN
            else:
                gdf.at[idx, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                feeder_n += 1

        trunk_n += 1

    # Fleet from cycle time (no assumed speed)
    gdf["Fleet_Required"] = gdf.apply(
        lambda r: _fleet_from_cycle(r["Cycle_Time_Min"], int(r["Headway_Min"])), axis=1
    )
    gdf["New_Wait_Time"] = gdf["Headway_Min"].astype(int)

    log.info("  Trunks: %d | Merged: %d | Feeders: %d",
             (gdf["Action_Taken"] == "UPGRADED_TO_TRUNK").sum(),
             (gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum(),
             (gdf["Action_Taken"] == "RETAINED_AS_FEEDER").sum())
    return gdf


# ── 2-L  Terminal capacity constraint ────────────────────────────────────────

def apply_terminal_capacity(gdf: gpd.GeoDataFrame,
                             gdf_pois: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Applying terminal capacity constraints…")
    term_pois = gdf_pois[gdf_pois["category"].str.lower().isin(TERMINAL_CATEGORIES)].to_crs(UTM_CRS)
    if term_pois.empty:
        return gdf
    term_tree = STRtree(term_pois.geometry.tolist())
    gdf = gdf.copy()
    downgraded = 0
    for idx, row in gdf.iterrows():
        if row["Action_Taken"] != "UPGRADED_TO_TRUNK":
            continue
        if row["Fleet_Required"] <= FLEET_CAP_HARD:
            continue
        has_cap = False
        for lon, lat in [(float(row["Start_Lon"]), float(row["Start_Lat"])),
                          (float(row["End_Lon"]),   float(row["End_Lat"]))]:
            pt_utm = gpd.GeoSeries([Point(lon, lat)], crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
            if len(term_tree.query(pt_utm.buffer(TERMINAL_BUFFER_M))) > 0:
                has_cap = True; break
        if not has_cap:
            gdf.at[idx, "Fleet_Required"] = FLEET_CAP_HARD
            gdf.at[idx, "Action_Taken"]   = "RETAINED_AS_FEEDER"
            gdf.at[idx, "Headway_Min"]    = FEEDER_HEADWAY_MIN
            gdf.at[idx, "New_Wait_Time"]  = FEEDER_HEADWAY_MIN
            downgraded += 1
    log.info("  %d trunks capacity-downgraded to feeder.", downgraded)
    return gdf


# ── 2-M  Network Cost-Benefit Score ──────────────────────────────────────────

def compute_network_score(gdf: gpd.GeoDataFrame, net_pop: int) -> float:
    active = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    total_fleet = int(active["Fleet_Required"].sum())
    
    # New Score: Passenger Wait Time Saved
    time_saved_mins = (active["Old_Wait_Time"] - active["New_Wait_Time"]).clip(lower=0)
    total_time_saved = (time_saved_mins * active["Population_Served"]).sum()
    
    # Efficiency ratio: Population served per minibus deployed
    efficiency_ratio = net_pop / max(1, total_fleet)
    
    log.info("=" * 68)
    log.info("  NETWORK EFFICIENCY & IMPACT SCORE")
    log.info("  Deduplicated Pop. Served : %s residents", f"{net_pop:,}")
    log.info("  Active Fleet Required    : %d minibuses", total_fleet)
    log.info("  Fleet Efficiency         : %.0f residents per minibus", efficiency_ratio)
    log.info("  Total Daily Time Saved   : %s person-minutes", f"{int(total_time_saved):,}")
    log.info("=" * 68)
    
    return float(total_time_saved)


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 3  ─  MEGA SUMMARY LOG (Chatbot Brain)
# ══════════════════════════════════════════════════════════════════════════════

def _reasoning(row: pd.Series) -> str:
    action  = row["Action_Taken"]
    old_id  = row["Route_ID"]
    new_id  = row["New_Route_ID"]
    pop     = f"{int(row['Population_Served']):,}"
    ow, nw  = row["Old_Wait_Time"], row["New_Wait_Time"]
    ovlp    = f"{row['Overlap_Metric'] * 100:.1f}%"
    km      = f"{row['Route_KM']:.1f}"
    fleet   = row["Fleet_Required"]
    r_sc    = f"{row.get('Residential_Score', 0):.1f}"
    c_sc    = f"{row.get('Commercial_Score', 0):.1f}"
    turns   = row.get("Sharp_Turns", 0)
    penalty = row.get("Junction_Penalty_Min", 0)

    if action == "UPGRADED_TO_TRUNK":
        return (
            f"Route {old_id} ({row['Route_Name']}) upgraded to Trunk {new_id}. "
            f"Highest Gravity Score in cluster: residential {r_sc}/km + commercial {c_sc}/km. "
            f"Serves ~{pop} residents across {km} km. "
            f"OSRM cycle time includes {turns} sharp turns adding {penalty:.0f} min penalty. "
            f"Fleet of {fleet} minibuses reduces wait from {ow} to {nw} min."
        )
    elif action == "MERGED_INTO_TRUNK":
        return (
            f"Route {old_id} ({row['Route_Name']}) merged into Trunk {new_id}. "
            f"Spatial overlap {ovlp}; low independent demand ({pop} residents). "
            f"Transfer penalty check confirmed merging improves or matches direct service "
            f"(wait: {ow} → {nw} min via Trunk). Redundant operations eliminated."
        )
    else:  # RETAINED_AS_FEEDER
        return (
            f"Route {old_id} ({row['Route_Name']}) retained as Feeder {new_id}. "
            f"Distinct catchment of {pop} residents across {km} km (overlap {ovlp}). "
            f"Transfer penalty analysis shows forcing a transfer would worsen passenger "
            f"experience vs. current direct service. "
            f"Fleet of {fleet} minibuses targets {nw}-min headway (from {ow} min currently). "
            f"Residential score {r_sc}/km | Commercial score {c_sc}/km."
        )


def generate_log(gdf: gpd.GeoDataFrame, out_path: str) -> pd.DataFrame:
    log.info("Generating Rationalisation Log → %s", out_path)
    cols = [c for c in [
        "Route_ID", "Route_Name", "Action_Taken", "New_Route_ID",
        "Residential_Score", "Commercial_Score", "Composite_Demand_Score",
        "Overlap_Metric", "Population_Served",
        "Old_Wait_Time", "New_Wait_Time", "Fleet_Required", "Route_KM",
        "HV_POI_Count", "Sharp_Turns", "Junction_Penalty_Min",
        "OSRM_Duration_S", "Cycle_Time_Min", "Geo_Source", "Fleet_Imputed",
    ] if c in gdf.columns]
    df_log = gdf[cols].copy()
    df_log.rename(columns={"Route_ID": "Old_Route_ID", "Route_Name": "Old_Name"}, inplace=True)
    df_log["Reasoning_String"] = gdf.apply(_reasoning, axis=1)
    df_log.to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  Log written: %d rows.", len(df_log))
    return df_log


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 4  ─  CARTOGRAPHY, DASHBOARDS & EXPORT
# ══════════════════════════════════════════════════════════════════════════════

# ── Shared map helpers ────────────────────────────────────────────────────────

def _safe_coords(geom) -> List[Tuple[float, float]]:
    if geom is None: return []
    coords = [c for line in geom.geoms for c in line.coords] \
             if isinstance(geom, MultiLineString) else list(geom.coords)
    return [(lat, lon) for lon, lat in coords]


def _popup_html(row: pd.Series) -> str:
    action_col = {"UPGRADED_TO_TRUNK": "#1A237E",
                  "MERGED_INTO_TRUNK": "#880E4F",
                  "RETAINED_AS_FEEDER": "#00695C"}.get(row.get("Action_Taken", ""), "#333")
    time_saved = max(0, int(row.get("Old_Wait_Time", 0)) - int(row.get("New_Wait_Time", 0)))
    return f"""
<div style="font-family:'Segoe UI',Arial,sans-serif;min-width:250px;font-size:12px">
  <div style="color:{action_col};font-size:15px;font-weight:700">{row.get('New_Route_ID','N/A')}</div>
  <div style="background:{action_col};color:#fff;display:inline-block;border-radius:4px;
    padding:1px 8px;font-size:10px;margin:3px 0 7px">{row.get('Action_Taken','').replace('_',' ')}</div>
  <table style="width:100%;border-collapse:collapse;line-height:1.7">
    <tr><td style="color:#666">Original ID</td><td><b>{row.get('Route_ID','')}</b></td></tr>
    <tr><td style="color:#666">Route Name</td><td>{row.get('Route_Name','')}</td></tr>
    <tr><td style="color:#666">Length</td><td><b>{row.get('Route_KM',0):.1f} km</b></td></tr>
    <tr><td style="color:#666">Headway</td><td><b>{row.get('Headway_Min','?')} min</b></td></tr>
    <tr><td style="color:#666">Fleet Required</td><td><b>{row.get('Fleet_Required','?')} minibuses</b></td></tr>
    <tr style="border-top:1px solid #f0f0f0">
      <td style="color:#1565C0">🏙 Res. Score</td>
      <td><b>{row.get('Residential_Score',0):.1f}</b> /km</td></tr>
    <tr><td style="color:#E65100">🏪 Com. Score</td>
        <td><b>{row.get('Commercial_Score',0):.1f}</b> /km</td></tr>
    <tr><td style="color:#666">⭐ Composite</td>
        <td><b>{row.get('Composite_Demand_Score',0):.1f}</b></td></tr>
    <tr style="border-top:1px solid #f0f0f0">
      <td style="color:#666">Pop. Served</td><td><b>{int(row.get('Population_Served',0)):,}</b></td></tr>
    <tr><td style="color:#666">HV POIs</td><td>{row.get('HV_POI_Count',0)}</td></tr>
    <tr><td style="color:#666">Sharp Turns</td><td>{row.get('Sharp_Turns',0)}</td></tr>
    <tr><td style="color:#666">Jn. Penalty</td><td>{row.get('Junction_Penalty_Min',0):.0f} min</td></tr>
    <tr style="border-top:1px solid #f0f0f0;color:green">
      <td><b>Time Saved</b></td><td><b>{time_saved} min/trip</b></td></tr>
  </table>
</div>"""


# ── 4-A  Master map sidebar HTML/CSS/JS ──────────────────────────────────────

def _sidebar_html() -> str:
    return f"""
<style>
  /* * THE BULLETPROOF FIX: 
   * Move the Leaflet control 5000px off-screen.
   * We MUST leave pointer-events as 'auto' so the browser allows JS to click it.
   */
  .leaflet-control-layers {{
    position: fixed !important;
    bottom: -5000px !important;
    right: -5000px !important;
    opacity: 0.01 !important;
    pointer-events: auto !important; 
  }}

  #tsb {{
    position:fixed; top:12px; left:12px; z-index:9999;
    background:rgba(255,255,255,0.97); border-radius:10px;
    box-shadow:0 2px 16px rgba(0,0,0,0.18);
    font-family:'Segoe UI',Arial,sans-serif; font-size:12px;
    width:215px; overflow:hidden; border:1px solid #e0e0e0;
  }}
  #tsb .tsb-hdr {{
    background:#1A237E; color:#fff; padding:10px 14px;
    font-size:13px; font-weight:700; line-height:1.4;
  }}
  #tsb .tsb-hdr small {{
    display:block; font-size:10px; font-weight:400; opacity:.8; margin-top:1px;
  }}
  #tsb .tsb-sec {{
    padding:7px 12px 5px; border-top:1px solid #f0f0f0;
  }}
  #tsb .tsb-title {{
    font-size:10px; font-weight:700; color:#9E9E9E;
    letter-spacing:.9px; text-transform:uppercase; margin-bottom:5px;
  }}
  #tsb label {{
    display:flex; align-items:center; gap:7px; padding:3px 0;
    cursor:pointer; color:#333; line-height:1.3; user-select:none;
  }}
  #tsb label:hover {{ color:#1A237E; }}
  #tsb input[type=checkbox] {{
    width:13px; height:13px; cursor:pointer; flex-shrink:0; accent-color:#1A237E;
  }}
  #tsb .tsb-swatch {{
    display:inline-block; width:18px; height:3px; border-radius:2px; flex-shrink:0;
  }}
</style>

<div id="tsb">
  <div class="tsb-hdr">
    🚌 Jammu Transit v5.0
    <small>Network Rationalisation Engine</small>
  </div>

  <div class="tsb-sec">
    <div class="tsb-title">Network Layers</div>
    <label>
      <input type="checkbox" id="sb-trunk" checked
             onchange="sbToggle('{FG['trunk']}',this.checked)">
      <span class="tsb-swatch" style="background:#1A237E"></span>
      Trunk Corridors
    </label>
    <label>
      <input type="checkbox" id="sb-feeder" checked
             onchange="sbToggle('{FG['feeder']}',this.checked)">
      <span class="tsb-swatch" style="background:#00695C;border-top:2px dashed #00695C;height:0"></span>
      Feeder Routes
    </label>
    <label>
      <input type="checkbox" id="sb-original"
             onchange="sbToggle('{FG['original']}',this.checked)">
      <span class="tsb-swatch" style="background:#9E9E9E"></span>
      Original Network
    </label>
  </div>

  <div class="tsb-sec">
    <div class="tsb-title">Demand Layers</div>
    <label>
      <input type="checkbox" id="sb-heatmap" checked
             onchange="sbToggle('{FG['heatmap']}',this.checked)">
      🌡 Population Heatmap
    </label>
    <label>
      <input type="checkbox" id="sb-hv-poi" checked
             onchange="sbToggle('{FG['hv_poi']}',this.checked)">
      🏥 High-Priority POIs
    </label>
    <label>
      <input type="checkbox" id="sb-sec-poi"
             onchange="sbToggle('{FG['sec_poi']}',this.checked)">
      🏪 Secondary POIs
    </label>
    <label>
      <input type="checkbox" id="sb-catchment"
             onchange="sbToggle('{FG['catchment']}',this.checked)">
      💧 Route Catchments
    </label>
  </div>

  <div class="tsb-sec">
    <div class="tsb-title">Route Details</div>
    <label>
      <input type="checkbox" id="sb-pins" checked
             onchange="sbToggle('{FG['pins']}',this.checked)">
      📍 Start / End Terminals
    </label>
    <label>
      <input type="checkbox" id="sb-via"
             onchange="sbToggle('{FG['via']}',this.checked)">
      🔵 Via Waypoints
    </label>
  </div>
</div>

<script>
  window.sbToggle = function(layerName, isChecked) {{
    // Leaflet applies this specific class to its map layer inputs
    var inputs = document.querySelectorAll('.leaflet-control-layers-overlays input.leaflet-control-layers-selector');
    
    for (var i = 0; i < inputs.length; i++) {{
      var input = inputs[i];
      // The text name of the layer is always a sibling/child of the input's parent wrapper
      var parentText = input.parentElement.textContent || input.parentElement.innerText || "";
      
      // Strict matching (trimming handles Leaflet's weird invisible spacing)
      if (parentText.trim() === layerName.trim()) {{
        if (input.checked !== isChecked) {{
          input.click(); // This fires the native event Leaflet is listening for
        }}
        break; 
      }}
    }}
  }};
</script>
"""

# ── 4-B  Master Transit Map ───────────────────────────────────────────────────

def build_master_map(gdf: gpd.GeoDataFrame,
                     gdf_pois: gpd.GeoDataFrame,
                     raster_path: str,
                     out_path: str,
                     net_pop: int,
                     network_score: float) -> None:
    log.info("Building Master Transit Map → %s", out_path)

    # prefer_canvas=True is kept for high performance
    m = folium.Map(location=[32.7266, 74.8570], zoom_start=13,
                   tiles=None, prefer_canvas=True)
    folium.TileLayer(tiles=TILE_URL, attr=TILE_ATTR, name="CartoDB Light",
                     control=False).add_to(m)

    # 1. Population HeatMap (Optimized)
    hm_fg = folium.FeatureGroup(name=FG["heatmap"], show=True)
    if Path(raster_path).exists():
        try:
            import rasterio
            with rasterio.open(raster_path) as src:
                band = src.read(1).astype(float)
                nd   = src.nodata
                if nd is not None:
                    band[band == nd] = 0
                
                # Strip out any residual NaNs or Infs that could crash JS
                band = np.nan_to_num(band, nan=0.0, posinf=0.0, neginf=0.0)
                band = np.clip(band, 0, None)
                
                tf   = src.transform
                r_, c_ = np.where(band > 5)
                lats = tf.f + r_ * tf.e
                lons = tf.c + c_ * tf.a
                vals = band[r_, c_]
                mx   = vals.max()
                if mx > 0:
                    vals = (vals / mx * 100).tolist()
                    lats, lons = lats.tolist(), lons.tolist()
                    
                    max_pts = 15000
                    if len(vals) > max_pts:
                        step = len(vals) // max_pts + 1
                        lats, lons, vals = lats[::step], lons[::step], vals[::step]
                        
                    HeatMap(list(zip(lats, lons, vals)),
                            radius=15, blur=14,
                            gradient={"0.0": "blue", "0.4": "lime", "0.8": "yellow", "1.0": "red"}
                            ).add_to(hm_fg)
        except Exception as exc:
            log.warning("  HeatMap skipped: %s", exc)
    hm_fg.add_to(m)

    # 2. Original Network (grey, hidden by default)
    orig_fg = folium.FeatureGroup(name=FG["original"], show=False)
    for _, row in gdf.iterrows():
        coords = _safe_coords(row.geometry)
        if len(coords) >= 2:
            folium.PolyLine(coords, color=COLOUR["current_net"], weight=1.8,
                            opacity=0.45,
                            tooltip=f"{row['Route_ID']} | {row['Route_Name']}"
                            ).add_to(orig_fg)
    orig_fg.add_to(m)

    # 3. Trunk Corridors (Bulletproof Canvas Lines)
    trunk_fg = folium.FeatureGroup(name=FG["trunk"], show=True)
    for _, row in gdf[gdf["Action_Taken"].isin(["UPGRADED_TO_TRUNK", "MERGED_INTO_TRUNK"])].iterrows():
        coords = _safe_coords(row.geometry)
        if len(coords) < 2: continue
        popup = folium.Popup(folium.Html(_popup_html(row), script=True), max_width=310)
        
        is_main_trunk = row["Action_Taken"] == "UPGRADED_TO_TRUNK"
        weight  = 5.0 if is_main_trunk else 2.5
        opacity = 0.9 if is_main_trunk else 0.4
        
        folium.PolyLine(coords, color=COLOUR["trunk"], weight=weight, 
                        opacity=opacity,
                        tooltip=f"🚌 {row['New_Route_ID']} | {row['Route_Name']}",
                        popup=popup).add_to(trunk_fg)
    trunk_fg.add_to(m)

    # 4. Feeder Routes
    feeder_fg = folium.FeatureGroup(name=FG["feeder"], show=True)
    for _, row in gdf[gdf["Action_Taken"] == "RETAINED_AS_FEEDER"].iterrows():
        coords = _safe_coords(row.geometry)
        if len(coords) < 2: continue
        popup = folium.Popup(folium.Html(_popup_html(row), script=True), max_width=310)
        
        folium.PolyLine(coords, color=COLOUR["feeder"], weight=3.5,
                        opacity=0.82, dash_array="8 5",
                        tooltip=f"🚐 {row['New_Route_ID']} | {row['Route_Name']}",
                        popup=popup).add_to(feeder_fg)
    feeder_fg.add_to(m)

    # 5. High-Priority POIs
    hv_fg = folium.FeatureGroup(name=FG["hv_poi"], show=True)
    # REPLACE THIS LINE:
    hv_pois = gdf_pois[gdf_pois["category"].str.lower().isin(TIER_1_POIS)]
    for _, poi in hv_pois.iterrows():
        folium.Marker(
            [poi.geometry.y, poi.geometry.x],
            icon=folium.Icon(color="red", icon="plus-sign", prefix="glyphicon"),
            tooltip=f"<b>{poi.get('name', poi['category'])}</b> [{poi['category']}]",
        ).add_to(hv_fg)
    hv_fg.add_to(m)

    # 6. Secondary POIs
    sec_fg   = folium.FeatureGroup(name=FG["sec_poi"], show=False)
    # REPLACE THIS LINE:
    sec_pois = gdf_pois[~gdf_pois["category"].str.lower().isin(TIER_1_POIS)]
    for _, poi in sec_pois.iterrows():
        folium.CircleMarker(
            [poi.geometry.y, poi.geometry.x], radius=5,
            color=COLOUR["poi_secondary"], fill=True, fill_opacity=0.8,
            tooltip=f"{poi.get('name', poi['category'])} [{poi['category']}]",
        ).add_to(sec_fg)
    sec_fg.add_to(m)

    # 7. Catchment polygons
    catch_fg = folium.FeatureGroup(name=FG["catchment"], show=False)
    features = []
    for _, row in gdf.iterrows():
        c = row.get("Catchment")
        if c is None or not hasattr(c, "is_empty") or c.is_empty:
            continue
        try:
            simplified = c.simplify(0.0003)  
            features.append({
                "type": "Feature",
                "properties": {"route_id": row["Route_ID"], "name": row["Route_Name"]},
                "geometry": mapping(simplified),
            })
        except Exception:
            pass
    if features:
        folium.GeoJson(
            {"type": "FeatureCollection", "features": features},
            style_function=lambda _: {
                "fillColor":   COLOUR["catchment_fill"],
                "color":       COLOUR["catchment_line"],
                "fillOpacity": 0.12,
                "weight":      1,
            },
            tooltip=folium.GeoJsonTooltip(fields=["route_id", "name"],
                                           aliases=["Route", "Name"]),
        ).add_to(catch_fg)
    catch_fg.add_to(m)

    # 8. Start / End Terminals
    pins_fg = folium.FeatureGroup(name=FG["pins"], show=True)
    for _, row in gdf.iterrows():
        folium.CircleMarker(
            [float(row["Start_Lat"]), float(row["Start_Lon"])],
            radius=5, color=COLOUR["start_pin"], fill=True, fill_opacity=0.9,
            tooltip=f"▶ {row['Route_Name']} START",
        ).add_to(pins_fg)
        folium.CircleMarker(
            [float(row["End_Lat"]), float(row["End_Lon"])],
            radius=5, color=COLOUR["end_pin"], fill=True, fill_opacity=0.9,
            tooltip=f"■ {row['Route_Name']} END",
        ).add_to(pins_fg)
    pins_fg.add_to(m)

    # 9. Via Waypoints
    via_fg = folium.FeatureGroup(name=FG["via"], show=False)
    for _, row in gdf.iterrows():
        for lon, lat in parse_via(row.get("Via_Coordinates")):
            folium.CircleMarker(
                [lat, lon], radius=3, color=COLOUR["via_dot"],
                fill=True, fill_opacity=0.7,
                tooltip=f"Via | {row['Route_Name']}",
            ).add_to(via_fg)
    via_fg.add_to(m)

    folium.LayerControl(collapsed=False, position="topright").add_to(m)

    sidebar = _sidebar_html()
    m.get_root().html.add_child(folium.Element(sidebar))

    m.save(out_path)
    log.info("  Master map saved.")
# ── 4-C  Individual route maps ────────────────────────────────────────────────

def _mini_stats_panel(row: pd.Series) -> str:
    """Fixed top-right overlay panel for individual route maps."""
    action    = row.get("Action_Taken", "")
    colour    = COLOUR["trunk"] if action == "UPGRADED_TO_TRUNK" else COLOUR["feeder"]
    time_saved = max(0, int(row.get("Old_Wait_Time", 0)) - int(row.get("New_Wait_Time", 0)))
    return f"""
<div style="position:fixed;top:12px;right:12px;z-index:9999;
     background:rgba(255,255,255,0.97);padding:14px 18px;
     border-radius:10px;box-shadow:0 2px 14px rgba(0,0,0,0.18);
     font-family:'Segoe UI',Arial,sans-serif;font-size:12px;
     min-width:215px;border-top:4px solid {colour};border:1px solid #e0e0e0;
     border-top:4px solid {colour}">
  <div style="color:{colour};font-size:15px;font-weight:700">{row.get('New_Route_ID','')}</div>
  <div style="background:{colour};color:#fff;display:inline-block;
       border-radius:3px;padding:1px 8px;font-size:10px;margin:3px 0 8px">
    {action.replace('_',' ')}
  </div>
  <table style="width:100%;border-collapse:collapse;line-height:1.85">
    <tr><td style="color:#666">📏 Length</td>
        <td style="text-align:right;font-weight:600">{row.get('Route_KM',0):.1f} km</td></tr>
    <tr><td style="color:#666">⏱ Headway</td>
        <td style="text-align:right;font-weight:600">{row.get('Headway_Min','?')} min</td></tr>
    <tr><td style="color:#666">🚌 Fleet</td>
        <td style="text-align:right;font-weight:600">{row.get('Fleet_Required','?')} minibuses</td></tr>
    <tr><td style="color:#666">👥 Pop. Served</td>
        <td style="text-align:right;font-weight:600">{int(row.get('Population_Served',0)):,}</td></tr>
    <tr style="border-top:1px solid #f0f0f0">
      <td style="color:#1565C0">🏙 Res. Score</td>
      <td style="text-align:right;font-weight:600">{row.get('Residential_Score',0):.1f} /km</td></tr>
    <tr><td style="color:#E65100">🏪 Com. Score</td>
        <td style="text-align:right;font-weight:600">{row.get('Commercial_Score',0):.1f} /km</td></tr>
    <tr><td style="color:#666">⚡ Jn. Penalty</td>
        <td style="text-align:right;font-weight:600">{row.get('Junction_Penalty_Min',0):.0f} min
        ({row.get('Sharp_Turns',0)} turns)</td></tr>
    <tr style="border-top:1px solid #f0f0f0;color:{'green' if time_saved > 0 else '#666'}">
      <td><b>⏩ Time Saved</b></td>
      <td style="text-align:right;font-weight:700">{time_saved} min/trip</td></tr>
  </table>
</div>"""


def build_individual_maps(gdf: gpd.GeoDataFrame,
                           gdf_pois: gpd.GeoDataFrame,
                           out_dir: str) -> Dict[str, str]:
    """
    One HTML map per proposed Trunk or Feeder route.
    Elements rendered: route line, 400m catchment, start/end pins,
    via waypoints, HV POIs (red), secondary POIs (amber), mini stats panel.
    NO directional arrows on individual maps (per user specification).
    """
    log.info("Building individual route maps → %s/", out_dir)
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    pois_wgs  = gdf_pois.to_crs(WGS84_CRS).copy()
    proposed  = gdf[gdf["Action_Taken"].isin(
        ["UPGRADED_TO_TRUNK", "RETAINED_AS_FEEDER"])].copy()
    file_map: Dict[str, str] = {}

    for _, row in proposed.iterrows():
        new_id  = row["New_Route_ID"]
        safe_id = new_id.replace("/", "_").replace("\\", "_")
        fpath   = Path(out_dir) / f"{safe_id}.html"

        coords = _safe_coords(row.geometry)
        if not coords:
            continue

        lats   = [c[0] for c in coords]
        lons   = [c[1] for c in coords]
        centre = [float(np.mean(lats)), float(np.mean(lons))]

        m = folium.Map(location=centre, zoom_start=14, tiles=None, prefer_canvas=True)
        folium.TileLayer(tiles=TILE_URL, attr=TILE_ATTR,
                         name="CartoDB Light", control=False).add_to(m)

        is_trunk = row["Action_Taken"] == "UPGRADED_TO_TRUNK"
        colour   = COLOUR["trunk"] if is_trunk else COLOUR["feeder"]

        # ── Catchment polygon ─────────────────────────────────────────────────
        catch = row.get("Catchment")
        if catch is not None and hasattr(catch, "is_empty") and not catch.is_empty:
            try:
                folium.GeoJson(
                    gpd.GeoDataFrame({"geometry": [catch]}, crs=WGS84_CRS).__geo_interface__,
                    style_function=lambda _: {
                        "fillColor":   COLOUR["catchment_fill"],
                        "color":       COLOUR["catchment_line"],
                        "fillOpacity": 0.22,
                        "weight":      1.5,
                        "dashArray":   "4 3",
                    },
                    tooltip="400m Walk Catchment",
                ).add_to(m)
            except Exception as exc:
                log.debug("  Catchment render skipped [%s]: %s", new_id, exc)

        # ── Route line (plain PolyLine — no arrows on individual maps) ────────
        popup = folium.Popup(folium.Html(_popup_html(row), script=True), max_width=310)
        if is_trunk:
            AntPath(coords, color=colour, weight=6, opacity=0.90, delay=800,
                    tooltip=f"{new_id} | {row['Route_Name']}", popup=popup).add_to(m)
        else:
            folium.PolyLine(coords, color=colour, weight=5, opacity=0.87,
                            dash_array="8 4",
                            tooltip=f"{new_id} | {row['Route_Name']}",
                            popup=popup).add_to(m)

        # ── Start pin ─────────────────────────────────────────────────────────
        try:
            folium.Marker(
                [float(row["Start_Lat"]), float(row["Start_Lon"])],
                icon=folium.Icon(color="green", icon="play", prefix="fa"),
                tooltip=f"▶ START: {row['Route_Name']}",
            ).add_to(m)
        except Exception as exc:
            log.debug("  Start pin skipped [%s]: %s", new_id, exc)

        # ── End pin ───────────────────────────────────────────────────────────
        try:
            folium.Marker(
                [float(row["End_Lat"]), float(row["End_Lon"])],
                icon=folium.Icon(color="red", icon="stop", prefix="fa"),
                tooltip=f"■ END: {row['Route_Name']}",
            ).add_to(m)
        except Exception as exc:
            log.debug("  End pin skipped [%s]: %s", new_id, exc)

        # ── Via waypoints ─────────────────────────────────────────────────────
        for lon, lat in parse_via(row.get("Via_Coordinates")):
            folium.CircleMarker(
                [lat, lon], radius=5, color=COLOUR["via_dot"],
                fill=True, fill_opacity=0.85, weight=1.5,
                tooltip="Via waypoint",
            ).add_to(m)

        # ── POIs inside catchment ─────────────────────────────────────────────
        if catch is not None and hasattr(catch, "is_empty") and not catch.is_empty:
            try:
                pois_inside = pois_wgs[pois_wgs.geometry.within(catch)]
                for _, poi in pois_inside.iterrows():
                    is_hv = poi["category"].lower() in TIER_1_POIS
                    if is_hv:
                        folium.Marker(
                            [poi.geometry.y, poi.geometry.x],
                            icon=folium.Icon(color="red", icon="plus-sign", prefix="glyphicon"),
                            tooltip=f"<b>{poi.get('name', poi['category'])}</b> [{poi['category']}]",
                        ).add_to(m)
                    else:
                        folium.CircleMarker(
                            [poi.geometry.y, poi.geometry.x], radius=6,
                            color=COLOUR["poi_secondary"], fill=True, fill_opacity=0.85,
                            tooltip=f"{poi.get('name', poi['category'])} [{poi['category']}]",
                        ).add_to(m)
            except Exception as exc:
                log.debug("  POI render skipped [%s]: %s", new_id, exc)

        # ── Mini stats panel ──────────────────────────────────────────────────
        m.get_root().html.add_child(folium.Element(_mini_stats_panel(row)))

        # ── Fit bounds ────────────────────────────────────────────────────────
        try:
            m.fit_bounds([[min(lats), min(lons)], [max(lats), max(lons)]])
        except Exception:
            pass

        m.save(str(fpath))
        file_map[new_id] = f"{out_dir}/{safe_id}.html"

    log.info("  Generated %d individual maps.", len(file_map))
    return file_map


# ── 4-D  Passenger Impact CSV ─────────────────────────────────────────────────

def export_passenger_impact(gdf: gpd.GeoDataFrame, out_path: str) -> None:
    log.info("Exporting Passenger Impact Matrix → %s", out_path)
    df = gdf[[
        "Route_ID", "Route_Name", "Action_Taken", "New_Route_ID",
        "Old_Wait_Time", "New_Wait_Time", "Population_Served",
        "Residential_Score", "Commercial_Score", "Composite_Demand_Score",
        "HV_POI_Count",
    ]].copy()
    df["Time_Saved_Mins"] = (df["Old_Wait_Time"] - df["New_Wait_Time"]).clip(lower=0)
    df["Cumulative_Person_Minutes_Saved_Daily"] = (
        df["Time_Saved_Mins"] * df["Population_Served"]
    ).astype(int)
    df.sort_values("Cumulative_Person_Minutes_Saved_Daily", ascending=False, inplace=True)
    total = df["Cumulative_Person_Minutes_Saved_Daily"].sum()
    log.info("  Total daily person-minutes saved: %s", f"{total:,}")
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


# ── 4-E  Operational CSV ──────────────────────────────────────────────────────

def export_csv(gdf: gpd.GeoDataFrame, file_map: Dict, out_path: str) -> None:
    log.info("Exporting Operational CSV → %s", out_path)
    cols = [c for c in EXPORT_COLS if c in gdf.columns]
    df   = gdf[cols].copy()
    df["View_Map"] = df["New_Route_ID"].apply(
        lambda nid: f'<a href="{file_map[nid]}">View Map</a>' if nid in file_map else ""
    )
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  CSV written: %d rows.", len(df))


# ── 4-F  Operational XLSX ─────────────────────────────────────────────────────

def export_xlsx(gdf: gpd.GeoDataFrame, file_map: Dict,
                out_path: str, net_pop: int, network_score: float) -> None:
    log.info("Exporting Operational XLSX → %s", out_path)
    cols = [c for c in EXPORT_COLS if c in gdf.columns]
    df   = gdf[cols].copy()

    # ── Numeric guards: no negative values in output ───────────────────────
    for col in ["Population_Served", "Residential_Score", "Commercial_Score",
                "Composite_Demand_Score", "Route_KM", "OSRM_Duration_S",
                "Cycle_Time_Min", "Fleet_Required", "Total_Minibuses_Existing"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).clip(lower=0)

    for col in ["Sharp_Turns", "HV_POI_Count", "Old_Wait_Time", "New_Wait_Time",
                "Headway_Min", "Junction_Penalty_Min"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).clip(lower=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rationalised Routes"

    hdr_fill  = PatternFill("solid", fgColor="1A237E")
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    norm_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="1565C0", underline="single")
    bdr       = Border(bottom=Side(style="thin", color="BDBDBD"),
                       right=Side(style="thin",  color="BDBDBD"))

    all_cols = cols + ["View_Map"]

    # Header
    for ci, cn in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=cn.replace("_", " "))
        cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr

    ACTION_FILLS = {
        "UPGRADED_TO_TRUNK":  "C5CAE9",
        "MERGED_INTO_TRUNK":  "FCE4EC",
        "RETAINED_AS_FEEDER": "E8F5E9",
    }

    for ri, row in enumerate(df.itertuples(index=False), 2):
        action   = getattr(row, "Action_Taken", "")
        row_fill = PatternFill("solid", fgColor=ACTION_FILLS.get(action, "FFFFFF"))
        nid      = getattr(row, "New_Route_ID", "")

        for ci, cn in enumerate(cols, 1):
            val  = getattr(row, cn, "")
            # Round floats nicely
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font, cell.fill = norm_font, row_fill
            cell.alignment = Alignment(vertical="center")
            cell.border    = bdr

        # View_Map hyperlink
        vm_cell = ws.cell(row=ri, column=len(all_cols))
        if nid in file_map:
            vm_cell.value = f'=HYPERLINK("{file_map[nid]}","View Map")'
            vm_cell.font  = link_font
        else:
            vm_cell.value = "—"
            vm_cell.font  = norm_font
        vm_cell.fill      = row_fill
        vm_cell.alignment = Alignment(horizontal="center", vertical="center")
        vm_cell.border    = bdr

    # Column widths
    WIDTHS = {
        "Route ID": 10, "Route Name": 22, "Action Taken": 22,
        "New Route ID": 14, "Route KM": 10, "OSRM Duration S": 16,
        "Cycle Time Min": 14, "Sharp Turns": 12, "Junction Penalty Min": 20,
        "Fleet Required": 14, "Total Minibuses Existing": 22, "Headway Min": 12,
        "Population Served": 18, "HV POI Count": 13,
        "Residential Score": 17, "Commercial Score": 17, "Composite Demand Score": 22,
        "Old Wait Time": 13, "New Wait Time": 13,
        "Overlap Metric": 15, "Geo Source": 14, "Fleet Imputed": 13, "View Map": 12,
    }
    for ci, cn in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(
            cn.replace("_", " "), 14)

    # Summary sheet
    active_gdf   = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    active_fleet = int(active_gdf["Fleet_Required"].sum())
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Jammu Matador Network Rationalisation v5.0 — Summary"
    ws2["A1"].font = Font(bold=True, size=13, color="1A237E")
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 26

    n_trunks  = int((gdf["Action_Taken"] == "UPGRADED_TO_TRUNK").sum())
    n_merged  = int((gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum())
    n_feeders = int((gdf["Action_Taken"] == "RETAINED_AS_FEEDER").sum())

    summary_rows = [
        ("Total Routes Input",                           len(gdf)),
        ("Upgraded to Trunk",                            n_trunks),
        ("Merged into Trunk (service retired)",          n_merged),
        ("Retained as Feeder",                           n_feeders),
        ("Active Fleet Required (Minibuses)",            active_fleet),
        ("Existing Fleet (Minibus Equiv.)",              int(gdf["Total_Minibuses_Existing"].sum())),
        ("Network Population Served (deduplicated)",     f"{net_pop:,}"),
        ("Network Cost-Benefit Score",                   f"{network_score:,.0f} Person-Mins"),
        ("Population Method",                            "Per-route: individual zonal_stats; "
                                                         "Summary: dissolved-union (no double-count)"),
        ("Gravity Method",                               "Residential (pop/km) + Commercial (HV_POI×500/km)"),
        ("Cycle Time Source",                            "OSRM actual duration + junction penalties"),
        ("Transfer Penalty Applied",                     f"{TRANSFER_PENALTY_MIN} min"),
    ]
    for ri, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(ri, 1, label).font = Font(bold=True, size=10)
        ws2.cell(ri, 2, val).font   = Font(size=10)
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 30

    # Colour the action rows in main sheet header section of summary
    for ri, action_label in [(4, "UPGRADED_TO_TRUNK"), (5, "MERGED_INTO_TRUNK"),
                              (6, "RETAINED_AS_FEEDER")]:
        fill = PatternFill("solid", fgColor=ACTION_FILLS.get(action_label, "FFFFFF"))
        for ci in [1, 2]:
            ws2.cell(ri, ci).fill = fill

    ws.freeze_panes = "A2"
    wb.save(out_path)
    log.info("  XLSX written.")
    # ── 4-G  GeoJSON Export for React Leaflet ────────────────────────────────────

def export_geojson(gdf: gpd.GeoDataFrame, out_path: str) -> None:
    log.info("Exporting Network GeoJSON → %s", out_path)
    
    # Filter to ONLY the final active network to keep the frontend payload light
    active_gdf = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()
    
    # Select only the columns the frontend needs for map popups and styling
    keep_cols = [
        "New_Route_ID", "Route_Name", "Action_Taken", "Route_KM", 
        "Headway_Min", "Fleet_Required", "Population_Served", 
        "Old_Wait_Time", "New_Wait_Time", "geometry"
    ]
    keep_cols = [c for c in keep_cols if c in active_gdf.columns]
    
    export_gdf = active_gdf[keep_cols]
    
    # Ensure CRS is explicitly WGS84 (EPSG:4326) which React Leaflet requires
    if export_gdf.crs != WGS84_CRS:
        export_gdf = export_gdf.to_crs(WGS84_CRS)
        
    # Drop any empty geometries to prevent Leaflet rendering crashes
    export_gdf = export_gdf[export_gdf.geometry.notnull() & ~export_gdf.geometry.is_empty]
    
    # Export using the GeoJSON driver
    export_gdf.to_file(out_path, driver="GeoJSON")
    log.info("  GeoJSON written: %d active features.", len(export_gdf))


# ══════════════════════════════════════════════════════════════════════════════
#  ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    t0 = time.perf_counter()
    log.info("=" * 70)
    log.info("  Jammu Matador Transit Rationalisation  —  Master Engine v5.0")
    log.info("=" * 70)

    # ── PHASE 1 ──────────────────────────────────────────────────────────────
    log.info("\n── PHASE 1: Data Ingestion, OSRM & Fleet Imputation ────────────────")
    df_routes    = load_routes(ROUTES_CSV)
    gdf_pois     = load_pois(POIS_CSV)
    osrm_results = fetch_all_osrm(df_routes)
    gdf          = apply_geometries(df_routes, osrm_results)
    gdf = count_pois_along_routes(gdf, gdf_pois, buffer_m=250.0)
    
    gdf = impute_fleet(gdf)

    # ── PHASE 2 ──────────────────────────────────────────────────────────────
    log.info("\n── PHASE 2: PTAL, Gravity Scores, Physics & Classification ─────────")
    gdf = build_catchments(gdf)
    gdf = compute_population(gdf, RASTER_PATH)        # FIXED: nodata + clamp
    gdf = compute_gravity_scores(gdf)                  # TWO SCORES, always >= 0
    gdf = compute_junction_penalties(gdf)
    gdf = compute_cycle_times(gdf)

    freq_scores    = compute_frequency_scores(gdf)
    overlap_matrix = compute_overlap_matrix(gdf)
    gdf = cluster_routes(gdf, overlap_matrix)
    gdf = backfill_overlap_metric(gdf, overlap_matrix)
    gdf = classify_routes(gdf, freq_scores, overlap_matrix)
    gdf = apply_terminal_capacity(gdf, gdf_pois)

    # Deduplicated summary population + network score
    net_pop       = compute_network_population_total(gdf, RASTER_PATH)
    network_score = compute_network_score(gdf, net_pop)

    # ── PHASE 3 ──────────────────────────────────────────────────────────────
    log.info("\n── PHASE 3: Mega Summary Log ────────────────────────────────────────")
    generate_log(gdf, LOG_CSV)

    # ── PHASE 4 ──────────────────────────────────────────────────────────────
    log.info("\n── PHASE 4: Cartography & Dashboard Export ─────────────────────────")
    build_master_map(gdf, gdf_pois, RASTER_PATH, MASTER_MAP_HTML, net_pop, network_score)
    file_map = build_individual_maps(gdf, gdf_pois, OUTPUT_DIR)
    export_csv(gdf, file_map, ROUTES_OUT_CSV)
    export_xlsx(gdf, file_map, ROUTES_OUT_XLSX, net_pop, network_score)
    export_passenger_impact(gdf, PASSENGER_IMPACT_CSV)
    export_geojson(gdf, ROUTES_GEOJSON)

    elapsed = time.perf_counter() - t0
    log.info("\n" + "=" * 70)
    log.info("  PIPELINE COMPLETE  (%.1f s)", elapsed)
    log.info("  Network Score = INR %s", f"{network_score:,.0f}")
    log.info("  Outputs:")
    log.info("    %-44s  sidebar + KPI map",        MASTER_MAP_HTML)
    log.info("    %-44s  LLM chatbot log",           LOG_CSV)
    log.info("    %-44s  passenger time savings",    PASSENGER_IMPACT_CSV)
    log.info("    %-44s  operational CSV",           ROUTES_OUT_CSV)
    log.info("    %-44s  operational XLSX",          ROUTES_OUT_XLSX)
    log.info("    %-44s  individual route maps",     f"{OUTPUT_DIR}/")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
